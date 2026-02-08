"""
SD-WAN Fleet Audit — Database Population (RAG data source)
============================================================
Parses all reference data (PDFs, Excel, datasheets) and stores everything
in SQLite so the LLM can query it. This is the "Retrieval" part of RAG.

Tables created/populated:
  - parc_actuel           (already exists — enriched with Excel data)
  - scenarios_migration   (filled by audit engine)
  - catalogue_reference   (already exists)
  - lifecycle             (NEW — EoS/EoL dates per model)
  - edge_7x0_specs        (NEW — performance specs of target models)
  - software_compatibility(NEW — which software runs on which model)
  - upgrade_paths         (NEW — step-by-step upgrade sequences)
  - mesures_detaillees    (NEW — full measured values per device from Excel)
"""

import sys
import sqlite3
from pathlib import Path

_SCRIPT_DIR = Path(__file__).parent            # src/
_PROJECT_ROOT = _SCRIPT_DIR.parent             # project root
sys.path.insert(0, str(_SCRIPT_DIR))

from audit_engine import DeviceRecord, audit_fleet, format_summary, format_detail_table

_DB_PATH = _PROJECT_ROOT / "data" / "hackathon_sdwan_v2.db"
_EXCEL_PATH = _PROJECT_ROOT / "data" / "data_hackathon_extended.xlsx"


# ---------------------------------------------------------------------------
# Reference data (extracted from official Arista PDFs)
# ---------------------------------------------------------------------------

LIFECYCLE_DATA = (
    # (model, eos_date, eol_date, urgency, status, recommended_replacement)
    ("Edge 500",  "2016-12-10", "2021-12-10", "CRITICAL", "Past EoL — aucun support vendeur",         "Edge 710"),
    ("Edge 510",  "2024-08-01", "2029-08-01", "HIGH",     "EoS atteint — planifier remplacement",     "Edge 710"),
    ("Edge 520",  "2021-03-25", "2027-03-25", "HIGH",     "EoS atteint — planifier remplacement",     "Edge 710"),
    ("Edge 520v", "2021-03-25", "2027-03-25", "HIGH",     "EoS atteint — planifier remplacement",     "Edge 720"),
    ("Edge 540",  "2021-03-25", "2027-03-25", "HIGH",     "EoS atteint — planifier remplacement",     "Edge 720"),
    ("Edge 610",  "2024-08-01", "2029-08-01", "HIGH",     "EoS atteint — planifier remplacement",     "Edge 710"),
    ("Edge 620",  "2025-04-01", "2030-04-01", "HIGH",     "EoS atteint — planifier remplacement",     "Edge 720"),
    ("Edge 640",  "2022-07-29", "2027-07-29", "HIGH",     "EoS atteint — planifier remplacement",     "Edge 740"),
    ("Edge 680",  "2022-07-29", "2027-07-29", "HIGH",     "EoS atteint — planifier remplacement",     "Edge 740"),
    ("Edge 840",  "2020-09-29", "2025-09-29", "CRITICAL", "Past EoL — aucun support vendeur",         "Edge 740"),
    ("Edge 1000", "2017-07-16", "2022-07-16", "CRITICAL", "Past EoL — aucun support vendeur",         "Edge 740"),
    ("Edge 2000", "2020-08-17", "2025-08-17", "CRITICAL", "Past EoL — aucun support vendeur",         "Edge 740"),
)

EDGE_7X0_SPECS = (
    # (model, throughput_imix_mbps, throughput_1300b_mbps, throughput_efw_mbps,
    #  max_tunnels, flows_per_sec, concurrent_flows, concurrent_flows_efw,
    #  max_nat_entries, sfp_slots, sfp_type, rj45_ports, rj45_speed,
    #  ram_gb, bw_tiers)
    ("Edge 710",  395,  950,  280,   50,  4000,  225000, 110000, 225000,  1, "1G SFP",     4, "1G",   4,  "10M,30M,50M,100M,200M,500M"),
    ("Edge 720", 2300, 5200, 1200,  400, 18000,  440000, 220000, 440000,  2, "1G/10G SFP+", 6, "2.5G", 8,  "10M,30M,50M,100M,200M,500M,1G,2G,10G"),
    ("Edge 740", 3500, 7200, 2000,  800, 26000,  900000, 450000, 900000,  2, "1G/10G SFP+", 6, "2.5G", 16, "100M,200M,500M,1G,2G,10G"),
)

SOFTWARE_COMPAT = (
    # (model, version_branch, supported, notes)
    ("Edge 840",  "4.2.x", "Oui", "Version actuelle du parc"),
    ("Edge 840",  "4.5.x", "Oui", "Intermediate upgrade step"),
    ("Edge 840",  "5.0.x", "Oui", "Supported"),
    ("Edge 840",  "5.2.x", "Oui", "Max version supportee pour Edge 840"),
    ("Edge 840",  "5.4.x", "Non", "Edge 840 non supporte apres 5.2.x"),
    ("Edge 840",  "6.x",   "Non", "Edge 840 ne supporte PAS la v6"),
    ("Edge 680",  "4.5.x", "Oui", "Supported"),
    ("Edge 680",  "5.0.x", "Oui", "Version actuelle du parc (5.0.0)"),
    ("Edge 680",  "5.2.x", "Oui", "Supported"),
    ("Edge 680",  "5.4.x", "Oui", "Supported"),
    ("Edge 680",  "6.1.x", "Oui", "Supported — LTS"),
    ("Edge 680",  "6.4.x", "Non", "Edge 680 EoS — probablement non supporte"),
    ("Edge 710",  "5.2.2+","Oui", "Minimum version pour Edge 7x0"),
    ("Edge 710",  "6.0.x", "Oui", "Supported"),
    ("Edge 710",  "6.1.x", "Oui", "Supported — LTS"),
    ("Edge 710",  "6.2.x", "Oui", "Supported"),
    ("Edge 710",  "6.4.x", "Oui", "Version cible — LTS candidate"),
    ("Edge 720",  "5.2.2+","Oui", "Minimum version pour Edge 7x0"),
    ("Edge 720",  "6.0.x", "Oui", "Supported"),
    ("Edge 720",  "6.1.x", "Oui", "Supported — LTS"),
    ("Edge 720",  "6.2.x", "Oui", "Supported"),
    ("Edge 720",  "6.4.x", "Oui", "Version cible — LTS candidate"),
    ("Edge 740",  "5.2.2+","Oui", "Minimum version pour Edge 7x0"),
    ("Edge 740",  "6.0.x", "Oui", "Supported"),
    ("Edge 740",  "6.1.x", "Oui", "Supported — LTS"),
    ("Edge 740",  "6.2.x", "Oui", "Supported"),
    ("Edge 740",  "6.4.x", "Oui", "Version cible — LTS candidate"),
)

UPGRADE_PATHS = (
    # (from_version, to_version, steps, order_notes)
    # Sources RN : 5.2.3 accepte 4.x+, 6.4.0 accepte 4.5.x+
    ("4.2.2", "6.4.x",
     "4.2.2 → 5.2.3 (LTS) → 6.1.x (LTS) → 6.4.x",
     "Ordre obligatoire : VCO → Gateways → Edges. "
     "Edge 840 : max 5.2.3 puis remplacement HW. Edge 680 : max 6.1.x puis remplacement HW."),
    ("4.5.x", "6.4.x",
     "4.5.x → 5.2.3 (LTS) → 6.1.x (LTS) → 6.4.x",
     "Ordre obligatoire : VCO → Gateways → Edges."),
    ("5.0.0", "6.4.x",
     "5.0.x → 6.4.x (direct, RN 6.4.0 accepte 4.5+)",
     "Ordre obligatoire : VCO → Gateways → Edges. "
     "Edge 680 : max 6.1.x puis remplacement HW."),
    ("5.2.x", "6.4.x",
     "5.2.x → 6.4.x (direct)",
     "Upgrade direct possible. VCO → Gateways → Edges."),
    ("5.4.x", "6.4.x",
     "5.4.x → 6.4.x (direct)",
     "Upgrade direct possible. VCO → Gateways → Edges."),
    ("6.1.x", "6.4.x",
     "6.1.x → 6.4.x (direct)",
     "Upgrade direct possible. VCO → Gateways → Edges."),
)


# ---------------------------------------------------------------------------
# Data loading (same as before)
# ---------------------------------------------------------------------------

def load_devices_from_db(db_path):
    conn = sqlite3.connect(db_path)
    try:
        cursor = conn.cursor()
        cursor.execute("""
            SELECT nom_hote, modele_actuel, version_logicielle,
                   debit_max_mesure, nb_ports_sfp_utilises, nb_tunnels_max
            FROM parc_actuel ORDER BY nom_hote
        """)
        rows = cursor.fetchall()
    finally:
        conn.close()
    return tuple(
        DeviceRecord(r[0], r[1], r[2], r[3] or 0, r[4] or 0, 0, r[5] or 0, 0, 0, 0)
        for r in rows
    )


def load_devices_from_excel(excel_path):
    try:
        import openpyxl
    except ImportError:
        print("WARNING: openpyxl not installed.")
        return ()
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb["Inventaire"]
    devices = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if not row[0]:
            continue
        dtype = str(row[3]).strip() if row[3] else ""
        if dtype and dtype.lower() != "edge":
            continue
        devices.append(DeviceRecord(
            str(row[0]).strip(),
            str(row[2]).strip() if row[2] else "Unknown",
            str(row[4]).strip() if row[4] else "Unknown",
            int(row[5]) if row[5] else 0,
            int(row[11]) if row[11] else 0,
            int(row[10]) if row[10] else 0,
            int(row[6]) if row[6] else 0,
            int(row[7]) if row[7] else 0,
            int(row[8]) if row[8] else 0,
            int(row[9]) if row[9] else 0,
        ))
    wb.close()
    return tuple(devices)


def merge_sources(db_devs, xl_devs):
    if not xl_devs:
        return db_devs
    xl_map = {d.hostname: d for d in xl_devs}
    merged, seen = [], set()
    for d in db_devs:
        merged.append(xl_map.get(d.hostname, d))
        seen.add(d.hostname)
    for d in xl_devs:
        if d.hostname not in seen:
            merged.append(d)
            seen.add(d.hostname)
    return tuple(sorted(merged, key=lambda x: x.hostname))


# ---------------------------------------------------------------------------
# Create and populate ALL reference tables
# ---------------------------------------------------------------------------

def create_reference_tables(db_path):
    """Create the reference tables needed for RAG."""
    conn = sqlite3.connect(db_path)
    c = conn.cursor()

    # Lifecycle table
    c.execute("DROP TABLE IF EXISTS lifecycle")
    c.execute("""
        CREATE TABLE lifecycle (
            modele TEXT PRIMARY KEY,
            date_fin_vente TEXT,
            date_fin_support TEXT,
            urgence TEXT,
            statut TEXT,
            remplacement_recommande TEXT
        )
    """)
    c.executemany(
        "INSERT INTO lifecycle VALUES (?,?,?,?,?,?)",
        LIFECYCLE_DATA,
    )

    # Edge 7x0 specs
    c.execute("DROP TABLE IF EXISTS edge_7x0_specs")
    c.execute("""
        CREATE TABLE edge_7x0_specs (
            modele TEXT PRIMARY KEY,
            debit_max_imix_mbps INTEGER,
            debit_max_1300b_mbps INTEGER,
            debit_max_firewall_mbps INTEGER,
            max_tunnels INTEGER,
            flows_par_seconde INTEGER,
            max_flux_concurrents INTEGER,
            max_flux_concurrents_firewall INTEGER,
            max_entrees_nat INTEGER,
            nb_ports_sfp INTEGER,
            type_sfp TEXT,
            nb_ports_rj45 INTEGER,
            vitesse_rj45 TEXT,
            ram_go INTEGER,
            tiers_bande_passante TEXT
        )
    """)
    c.executemany(
        "INSERT INTO edge_7x0_specs VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
        EDGE_7X0_SPECS,
    )

    # Software compatibility matrix
    c.execute("DROP TABLE IF EXISTS software_compatibility")
    c.execute("""
        CREATE TABLE software_compatibility (
            modele TEXT,
            branche_version TEXT,
            supporte TEXT,
            notes TEXT
        )
    """)
    c.executemany(
        "INSERT INTO software_compatibility VALUES (?,?,?,?)",
        SOFTWARE_COMPAT,
    )

    # Upgrade paths
    c.execute("DROP TABLE IF EXISTS upgrade_paths")
    c.execute("""
        CREATE TABLE upgrade_paths (
            version_source TEXT,
            version_cible TEXT,
            etapes TEXT,
            notes_ordre TEXT
        )
    """)
    c.executemany(
        "INSERT INTO upgrade_paths VALUES (?,?,?,?)",
        UPGRADE_PATHS,
    )

    # Detailed measured values per device (from Excel)
    c.execute("DROP TABLE IF EXISTS mesures_detaillees")
    c.execute("""
        CREATE TABLE mesures_detaillees (
            nom_hote TEXT PRIMARY KEY,
            modele TEXT,
            version TEXT,
            debit_max_mbps INTEGER,
            max_tunnels INTEGER,
            max_flows_par_sec INTEGER,
            max_flux_concurrents INTEGER,
            max_entrees_nat INTEGER,
            ports_rj45_utilises INTEGER,
            ports_sfp_utilises INTEGER
        )
    """)

    conn.commit()
    conn.close()


def update_parc_actuel(db_path, devices):
    """Update parc_actuel with correct model/version from merged data.

    The original DB has all 90 devices as 'Edge 840' / '4.2.2' but the
    Excel source has the correct Edge 680 entries with version 5.0.0.
    This function corrects the DB to match reality.
    """
    conn = sqlite3.connect(db_path)
    c = conn.cursor()
    for d in devices:
        c.execute(
            """UPDATE parc_actuel
               SET modele_actuel = ?, version_logicielle = ?,
                   debit_max_mesure = ?, nb_ports_sfp_utilises = ?,
                   nb_tunnels_max = ?
               WHERE nom_hote = ?""",
            (d.model, d.version, d.throughput_mbps, d.sfp_ports,
             d.tunnels, d.hostname),
        )
    conn.commit()
    conn.close()


def populate_measured_values(db_path, devices):
    """Fill mesures_detaillees from merged device data."""
    conn = sqlite3.connect(db_path)
    c = conn.cursor()
    c.execute("DELETE FROM mesures_detaillees")
    for d in devices:
        c.execute(
            "INSERT INTO mesures_detaillees VALUES (?,?,?,?,?,?,?,?,?,?)",
            (d.hostname, d.model, d.version, d.throughput_mbps, d.tunnels,
             d.flows_per_sec, d.concurrent_flows, d.nat_entries,
             d.rj45_ports, d.sfp_ports),
        )
    conn.commit()
    conn.close()


def populate_scenarios(db_path, results):
    conn = sqlite3.connect(db_path)
    c = conn.cursor()
    c.execute("DELETE FROM scenarios_migration")
    for r in results:
        c.execute("""
            INSERT INTO scenarios_migration
                (nom_hote, nouveau_modele_suggere, nouvelle_licence_suggeree,
                 chemin_migration_logicielle, cout_total_estime, complexite_intervention)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (r.device.hostname, r.target.target_model, r.target.license_tier,
              r.target.upgrade_path, r.target.cost, r.target.complexity))
    conn.commit()
    c.execute("SELECT COUNT(*) FROM scenarios_migration")
    count = c.fetchone()[0]
    conn.close()
    return count


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    db = str(_DB_PATH)
    xl = str(_EXCEL_PATH)

    print("=" * 70)
    print("SD-WAN Fleet Audit — Population de la base RAG")
    print("=" * 70)
    print()

    # 1. Load devices
    print("[1/6] Chargement de l'inventaire...")
    db_devs = load_devices_from_db(db)
    xl_devs = load_devices_from_excel(xl)
    devices = merge_sources(db_devs, xl_devs)
    print(f"      {len(devices)} devices charges (DB={len(db_devs)}, Excel={len(xl_devs)})")

    # 2. Create reference tables
    print("[2/6] Creation des tables de reference (lifecycle, specs, compatibilite)...")
    create_reference_tables(db)
    print("      Tables creees : lifecycle, edge_7x0_specs, software_compatibility, upgrade_paths, mesures_detaillees")

    # 3. Update parc_actuel with correct models/versions from Excel
    print("[3/6] Mise a jour de parc_actuel (correction modeles/versions)...")
    update_parc_actuel(db, devices)
    print(f"      parc_actuel corrige (Edge 680 + versions)")

    # 4. Populate measured values
    print("[4/6] Insertion des mesures detaillees...")
    populate_measured_values(db, devices)
    print(f"      {len(devices)} lignes dans mesures_detaillees")

    # 5. Run audit engine
    print("[5/6] Execution du moteur d'audit...")
    results = audit_fleet(devices)
    count = populate_scenarios(db, results)
    print(f"      {count} scenarios inseres dans scenarios_migration")

    # 6. Summary
    print("[6/6] Resume...")
    print()
    print(format_summary(results))
    print()
    print(format_detail_table(results))

    # Verify all tables
    print()
    print("=" * 70)
    print("VERIFICATION — Contenu de la base RAG")
    print("=" * 70)
    conn = sqlite3.connect(db)
    c = conn.cursor()
    for table in ("parc_actuel", "scenarios_migration", "catalogue_reference",
                   "lifecycle", "edge_7x0_specs", "software_compatibility",
                   "upgrade_paths", "mesures_detaillees"):
        c.execute(f"SELECT COUNT(*) FROM {table}")
        print(f"  {table}: {c.fetchone()[0]} lignes")
    conn.close()
    print()

    return results


if __name__ == "__main__":
    main()
